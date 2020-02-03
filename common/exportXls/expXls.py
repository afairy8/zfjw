import openpyxl as opl
import os
import uuid

defaultPath='D:\\projects\\jwAuto\\common\\expfiles'

def getfilename(filename=''):
    '''处理文件名
    :param filename:如果不包含路径，则默认路径，如果不指定文件名，则只用默认文件名，如果两者都不指定，那么都是默认
    :param content: 写入xls的内容，为二级嵌套格式，如[[]],或[()]
    :return:filename
    '''
    base=os.path.split(filename)
    if base[0] and base[1]:
        path=base[0]
        if not base[1].endswith('.xlsx'):
            name=base[1]+'.xlsx'
        else:
            name=base[1]
    elif base[0] and not base[1]:
        path=base[0]
        name=str(uuid.uuid1()).replace('-','')+'.xlsx'
    elif not base[0] and base[1]:
        path=defaultPath
        if not base[1].endswith('.xlsx'):
            name=base[1]+'.xlsx'
        else:
            name=base[1]
    else:
        path=defaultPath
        name=str(uuid.uuid1()).replace('-','')+'.xlsx'
    pathname=os.path.join(path,name)
    return pathname
def exp(filename,content=[],mode='new'):
    '''
    :param filename:如果不包含路径，则默认路径，如果不指定文件名，则只用默认文件名，如果两者都不指定，那么都是默认
    :param content: 写入xls的内容，为二级嵌套格式，如[[]],或[()]
    :return:filename
'''
    filename=getfilename(filename)
    if os.path.exists(filename):
        if mode=='new':
            filename=filename+str(uuid.uuid1()).replace('-','')+'.xlsx'
            wb=opl.Workbook()
        else:
            wb=opl.load_workbook(filename)
    else:
        wb=opl.Workbook()
    ws=wb.active
    for data in content:
        ws.append(data)
    wb.save(filename)
    return filename

