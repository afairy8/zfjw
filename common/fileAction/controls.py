import os
import time
import win32com.client as win32
import zipfile
import uuid
from common.fileAction import vars
import openpyxl as opl

class fileInfo:
    def __init__(self, fileName):
        self.fileName = fileName
        self._defaultPath=self.setDefaultPath()#defaultPath()
    def setDefaultPath(self):
        '''默认保存路径'''
        if not os.path.exists(vars.defaultPath):
            os.mkdir(vars.defaultPath)
        return vars.defaultPath
    def isExists(self):
        path,name=os.path.split(self.fileName)
        if not name:
            sign=0
        else:
            if not path:
                path=self._defaultPath
            sign=os.path.exists(os.path.join(path,name))
        # print(sign)
        return sign

    def getFileInfo(self):
        '''文件的创建时间'''
        if self.isExists():
            creatTime = time.strftime('%Y-%m-%d', time.gmtime(os.path.getctime(self.fileName)))
            return {'createTime': creatTime}

    def removeFile(self):
        '''删除文件'''
        if self.isExists():
            os.remove(self.fileName)
            return 1

    def xlsToXlsx(self, mode='new'):
        '''从xls文件转换为xlsx文件'''
        if self.isExists() and self.fileName.lower().strip().endswith('xls'):
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(self.fileName)
            if os.path.exists(self.fileName + 'x'):
                if mode.lower().strip() == 'new':
                    os.remove(self.fileName + 'x')
                    wb.SaveAs(self.fileName+'x', FileFormat=51)
                    wb.Close()
                    excel.Application.Quit()
            else:
                wb.SaveAs(self.fileName + 'x', FileFormat=51)
                wb.Close()
                excel.Application.Quit()
        else:
            pass
    def unzip(self,mode='new'):
        '''如果是压缩文件，则解压该文件'''
        if self.isExists() and zipfile.is_zipfile(self.fileName):
            zip=zipfile.ZipFile(self.fileName,'r')
            ziplist=zip.namelist()
            for z in ziplist:
                if os.path.exists(os.path.join(os.getcwd(),z)):
                    if mode=='new':
                        os.remove(os.path.join(os.getcwd(),z))
                        zip.extract(z)
                else:
                    zip.extract(z)
    def expFilename(self,suffix='.xlsx'):
        '''处理文件名
        :param filename:如果不包含路径，则默认路径，如果不指定文件名，则只用默认文件名，如果两者都不指定，那么都是默认
        :param content: 写入xls的内容，为二级嵌套格式，如[[]],或[()]
        :return:filename
        '''
        base = os.path.split(self.fileName)
        if base[0] and base[1]:
            path = base[0]
            if not base[1].endswith(suffix):
                name = base[1] + suffix
            else:
                name = base[1]
        elif base[0] and not base[1]:
            path = base[0]
            name = str(uuid.uuid1()).replace('-', '') + suffix
        elif not base[0] and base[1]:
            path = self._defaultPath
            if not base[1].endswith(suffix):
                name = base[1] + suffix
            else:
                name = base[1]
        else:
            path = self._defaultPath
            name = str(uuid.uuid1()).replace('-', '') + suffix
        pathname = os.path.join(path, name)
        return pathname
    def expXlsx(self, content=[], mode='new',suffix='.xlsx',sheetName='Sheet'):
        '''将content写入到指定的SheetName中
    '''
        xlsxFileName=self.expFilename(suffix)
        if suffix=='.xlsx':
            if os.path.exists(xlsxFileName):
                if mode=='new':
                    xlsxFileName=xlsxFileName+str(uuid.uuid1()).replace('-', '') + suffix
                    wb=opl.Workbook()
                else:
                    try:
                        wb=opl.load_workbook(xlsxFileName)
                    except:
                        wb=opl.Workbook()
            else:
                wb=opl.Workbook()
            wslist=wb.sheetnames
            if sheetName not in wslist:
                if sheetName is None:
                    sheetName='Sheet1'
                wb.create_sheet(sheetName)
            ws=wb[sheetName]
            ws.title=sheetName
            for data in content:
                ws.append(data)
            wb.save(xlsxFileName)
        elif suffix=='.txt':
            if os.path.exists(xlsxFileName):
                if mode=='new':
                    xlsxFileName=xlsxFileName+str(uuid.uuid1()).replace('-', '') + suffix
                    wb=open(xlsxFileName,'w',encoding='utf-8')
                else:
                    wb=open(xlsxFileName,'a',encoding='utf-8')
            else:
                wb=open(xlsxFileName,'w',encoding='utf-8')
            wb.write(content)
            wb.close()
        return 1
    def removeXlsxWs(self,sheetName='Sheet1'):
        '''self本身必须是xlsx格式,且存在'''
        if self.isExists() and self.fileName.endswith('.xlsx'):
            wb=opl.load_workbook(self.fileName)
            for name in wb.get_sheet_names:
                wb.remove(name)
            wb.create_sheet(sheetName)
    def getFileContent(self,sheetName='sheet1',type='active',containTitle=False):
        '''如果给定sheetName,则返回给定的SheetName的内容，如果没有找到相应的sheetName，如给定type='active'则
        返回活动工作表的内容，如不给定active则返回为空；
        如果不给定sheetName则返回所有工作表的内容！
        '''
        if self.isExists():
            if self.fileName.endswith('.xlsx'):
                wb=opl.load_workbook(self.fileName)
                wslist=wb.sheetnames
                L=[]
                ws=None
                if containTitle:
                    min_row=1
                else:
                    min_row=2
                if sheetName:
                    for name in wslist:
                        if name.lower().find(sheetName)>=0:
                            ws=wb[name]
                        else:
                            ws=None
                        if ws:
                            L.append(('内容来自于文件名={},工作表={}'.format(self.fileName,name),' '))
                            for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                rowL = []
                                for cell in row:
                                    if cell.value is None:
                                        rowL.append('')
                                    else:
                                        rowL.append(cell.value)
                                L.append(tuple(rowL))
                    if type=='active':
                        if ws is None and len(L)==0:
                            ws=wb.active
                            for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                rowL = []
                                for cell in row:
                                    if cell.value is None:
                                        rowL.append('')
                                    else:
                                        rowL.append(cell.value)
                                L.append(tuple(rowL))
                else:
                    for name in wslist:
                        ws=wb[name]
                        L.append(('内容来自于文件名={},工作表={}'.format(self.fileName, name), ' '))
                        for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            rowL = []
                            for cell in row:
                                if cell.value is None:
                                    rowL.append('')
                                else:
                                    rowL.append(cell.value)
                            L.append(tuple(rowL))
                if not len(L):
                    L=[('filename={}不存在指定的sheet={},且没有要求要返回active的工作表'.format(self.fileName,sheetName),'')]
                return L
            else:
                return [('格式不受支持','格式不受支持')]

def pathCommon(path,resdirs=[],resfiles=[]):
    '''返回根目录下的子目录与文件集合{'dirs':resdirs,'files':resfiles}'''
    if path:
        for root,dirs,files in os.walk(path):
            for file in files:
                resfiles.append(os.path.join(path,file))
            for dir in dirs:
                resdirs.append(os.path.join(path,dir))
                pathCommon(os.path.join(path,dir),resdirs,resfiles)
        return {'dirs':resdirs,'files':resfiles}



