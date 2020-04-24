import openpyxl as opl
from apps.jwglxt.cjgl.jhscjgl import vars
import os
from common.fileAction.controls import fileInfo
def init(con,xsFilePath):
    '''初始化课程库，成绩表'''

    for table in vars.tmpTableNames:
        res=con.objectExists(table)
        if res[0][0]:
            con.execute('drop table {}'.format(table))
    ###创建临时课程库、成绩库，并把临时课程库填满
    con.execute(vars.createJhsKck)
    print(vars.createJhsCj)
    con.execute(vars.createJhsCj)
    con.execute(vars.initJhsKck)
    ###将所有的xls文件转换为xlsx文件
    files=os.listdir(xsFilePath)
    for file in files:
        if file.endswith('.xls'):
            xlsx=fileInfo(os.path.join(xsFilePath,file))
            xlsx.xlsToXlsx()
        else:
            pass
    return 1


def genKcdm(con,kcmc,kcxz,xf,kcywmc):
    '''生成课程代码'''
    kcdmIsExists=con.execute(vars.kcdmIsExists.format(kcmc,xf,kcxz))
    #print(kcdmIsExists)
    if  kcdmIsExists:#存在则用现有的
        kcdm=kcdmIsExists[0][0]
    else:#不存在则新生成,并插入
        kcdm=con.execute(vars.kcdmMax)[0][0]
        L=[]
        L.append((kcdm,kcdm,kcmc,kcywmc,xf,kcxz))
        con.execute(vars.inJhsKck,L)
    #print(kcdm+kcmc)
    return kcdm

def uniCjxn(con):
    '''默认为当前学年的上一个学期'''
    dqxnxq=con.execute(vars.getDQXNXQ)[0]
    if dqxnxq[1]=='12':
        xn,xq=dqxnxq[0],'3'
    else:
        xn,xq=str(int(dqxnxq[0])-1),'12'
    return con.execute(vars.getCJXNXQ.format(xn,xq))[0]

def uniCj(cj):
    if cj[0:1] in ['优','良','中','及格','不及格']:
        fz='五级制'
    elif cj[0:1] in ['A','B','C','D','E','F']:
        fz='十一级制'
    else:
        fz='百分制'
    return fz

def uniKcxzGs(kcxz):
    '''课程性质课程归属'''
    xz=kcxz.split('-')[0] if len(kcxz.split('-'))>=1 else None
    gs=kcxz.split('-')[1] if len(kcxz.split('-'))>1 else None
    resxz=uniKcxz(xz)
    return (resxz,gs)

def uniKcxz(kcxz):
    if kcxz=="必修":
        return "专业必修课程"
    elif kcxz=="专选":
        return "专业选修课程"
    elif kcxz=="通选":
        return "通识类选修课程"
    else:
        return kcxz

def readXs(con,xsFilePath):
    '''读取学生成绩文件'''
    if init(con,xsFilePath):
        print('初始化成功，表likai_jhs_kck,likai_jhs_cj已创建，{}中的xls文件均已转换为xlsx格式'.format(xsFilePath))
    else:
        print('初始化失败')
    files=os.listdir(xsFilePath)
    kclist = []
    for file in files:
        if file.endswith('.xlsx'):
            wb=opl.load_workbook(os.path.join(xsFilePath,file))
            ws=wb.worksheets[0]
            if ws.title!='Sheet1':
                print('{}的文件模板有可能被修改'.format(file))
            xh,xm,cjxn,cjxq=None,None,None,None
            for row in ws.iter_rows(min_row=2,max_row=20,min_col=1,max_col=ws.max_column):
                if row[0].value:
                    if row[0].value.find('姓名')>=0:
                        xm=row[2].value
                    if row[0].value.find('学号')>=0:
                        xh=str(row[2].value).strip()
                    if row[0].value.find('成绩录入学年')>=0:
                        cjxn=row[2].value
                if row[4].value:
                    if row[4].value.find('成绩录入学期')>=0:
                        cjxq=str(row[8].value)
                if row[6].value and row[11].value:
                    if row[11].value and row[6].value.find('抵换')>=0:
                        kcmc=row[0].value
                        kcywmc=row[3].value.strip().replace('  ',' ').title()
                        kcxz=uniKcxzGs(row[7].value)[0]
                        xf=str(row[10].value)
                        cj=str(row[11].value).replace(' ','')
                        fz=uniCj(cj)
                        kch=genKcdm(con,kcmc,kcxz,xf,kcywmc)
                        if not cjxn or not cjxq:
                            cjxnxq=uniCjxn(con)
                            cjxn=cjxnxq[0]
                            cjxq=cjxnxq[1]
                        t=(cjxn,cjxq,fz,kch,kcxz,xh,cj,xm,uniKcxzGs(row[7].value)[1])
                        kclist.append(t)
    if kclist:
        con.execute(vars.inJhsCj,kclist)##将成绩写到成绩临时表
        con.execute(vars.writeToKck)##将课程写回课程库
        xlsx=fileInfo(os.path.join(xsFilePath,'当前文件夹的交换生成绩集合.xlsx'))
        content=[
            ('学年(必填)','学期(必填)','计分制(必填)','课程号(必填)','课程性质(必填)','成绩性质(必填)','学号(必填)'
             ,'成绩值(必填)','成绩备注','姓名','年级','班级','课程标记(必填)','课程类别','课程归属')
        ]
        res=con.execute('''select * from likai_jhc_cj''')
        content.extend(res)
        xlsx.expXlsx(content=content)
    return 1


def jhsCjInterface(con,xsFilePath):
    if xsFilePath and os.path.exists(xsFilePath):
        if readXs(con,xsFilePath=xsFilePath):
            return '{}路径下的补录成绩表已生成！'.format(xsFilePath)
    else:
        return '学生成绩文件路径明显有误！'