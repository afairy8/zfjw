from databaseconfig.connectdbs import connect
from apps.syjxxt import vars as syvar
import os
import openpyxl as opl
import subprocess as subp
from datetime import datetime

def restartService(serviceName):
    '''重启mssqlserver服务'''
    commad = 'runas /savecred /user:Administrator "sc stop {}"'.format(serviceName)
    stop = subp.Popen(commad)
    stop.wait(3)
    if stop.returncode == 0:
        print('{} 服务关闭成功'.format(serviceName))
    else:
        print('{} 服务关闭失败'.format(serviceName))
    commad = 'runas /savecred /user:Administrator "net start {}"'.format(serviceName)
    start = subp.Popen(commad)
    start.wait(10)
    if start.returncode == 0:
        print('{} 服务开启成功'.format(serviceName))
    else:
        print('{} 服务开启失败'.format(serviceName))
    return 1


def restoreMssqlDatabse(con):
    '''恢复实验教学系统的数据库,并且创建链路，将机构，专业，班级，学生、教师信息读入
    并且插入、更新、机构、专业、班级、学生、教师信息'''
    ###格式化恢复语句
    print('当前正执行机构，专业，班级，学生、教师信息同步')
    L = []
    dbs = ('lip_assets', 'lip_baseinfo', 'lip_custom', 'lip_teaching')
    files = os.listdir(syvar.sourcePath)
    for file in files:
        fi = file.strip().lower()
        for db in dbs:
            if fi.find(db) >= 0:
                restorCode = syvar.restoreCode.format(
                    db,
                    os.path.join(syvar.sourcePath, file),
                    db,
                    os.path.join(syvar.dbStorePath, db) + '.mdf',
                    db,
                    os.path.join(syvar.dbStorePath, db) + '.ldf',
                )
                L.append(restorCode)
    ####创建恢复脚本文件####
    if os.path.exists(os.path.join(os.getcwd(), 'resotre.sql')):
        os.remove(os.path.join(os.getcwd(), 'resotre.sql'))
    f = open(os.path.join(os.getcwd(), 'resotre.sql'), 'w')
    f.write(';\n'.join(L))
    f.write(syvar.addlink + '\n' + syvar.readJgZyBjXj + '\n' + syvar.insertbj)
    f.close()
    ##执行恢复脚本
    restoreCommad = syvar.sqlcmd.format(os.path.join(os.getcwd(), 'resotre.sql'))
    con.executeRestore(restoreCommad)
    con.commits()
    return 1


def drkc(con):
    '''导入课程'''
    print('当前正处理导入课程，创建的表为likai_new_course')
    createTmpTable = '''
    if object_id('lip_baseinfo.dbo.likai_new_course') is not null
        begin
        drop table lip_baseinfo.dbo.likai_new_course
        end
    create table lip_baseinfo.dbo.likai_new_course(
        kch varchar(255),
        kcmc varchar(255),
        kkdepid varchar(255),
        xs varchar(255),
        syzxs varchar(255),
        kcxz varchar(2),
        kkbm_dm varchar(255),
        kkxy varchar(255)
        )
    '''
    con.execute(createTmpTable)  ###创建临时课程信息表
    L = []
    wb = opl.load_workbook(os.path.join(syvar.sourcePath, 'sykc.xlsx'))
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        kkxy, kch, kcmc, syxs, xs, kcxz, kkbm_dm = \
            row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[1].value[2:4]
        if not syxs:
            syxs = xs
        if kcxz.find('非独立') >= 0:
            kcxz = '0'
        else:
            kcxz = '1'
        t = (kch, kcmc, syxs, xs, kcxz, kkbm_dm, kkxy)
        L.append(t)
    insertKcCode = '''
    insert into lip_baseinfo.dbo.likai_new_course(kch,kcmc,syzxs,xs,kcxz,kkbm_dm,kkxy)
    values(%s,%s,%s,%s,%s,%s,%s)
    '''
    con.execute(insertKcCode, L)  ###将课程插入临时表
    updateLikaiNewCourse = '''
    update lip_baseinfo.dbo.likai_new_course set kkdepid
=cast(t.id as varchar(255)) from lip_baseinfo.dbo.Base_Department t
where lip_baseinfo.dbo.likai_new_course.kkbm_dm=t.DepartmentCode;
    '''
    con.execute(updateLikaiNewCourse)  ##更新课程临时表中的信息
    insertBaseCourse = '''
    insert into lip_baseinfo.dbo.Base_CourseInfo(kch,kcmc,kkdepid,xs,syzxs,kcxz)
select kch,kcmc,kkdepid,xs,syzxs,kcxz from lip_baseinfo.dbo.likai_new_course t
where t.kch not in (select kch from lip_baseinfo.dbo.Base_CourseInfo)
    '''
    con.execute(insertBaseCourse)  ##将临时表中的课程信息插入到正式表
    return 1


def expRoom(con):
    '''导出实验房间信息'''
    if os.path.exists(os.path.join(os.getcwd(), '场地信息.xlsx')):
        os.remove(os.path.join(os.getcwd(), '场地信息.xlsx'))
    expSyCode = '''select id,roomname from lip_baseinfo.dbo.base_roominfo'''
    wb = opl.Workbook()
    ws = wb.create_sheet('sycd')
    ws.append(['id', 'roomname'])
    res = con.execute(expSyCode)
    for data in res:
        ws.append(data)
    expJwCode = '''
    select * from openquery(JWXT,'select cdbh,cdmc from JW_JCDM_CDXQXXB where XNM=''2019'' and xqm=''12'' and 1=1')
    '''
    ws = wb.create_sheet('jwcd')
    ws.append(['cdbh', 'cdmc'])
    res = con.execute(expJwCode)
    for data in res:
        ws.append(data)
    wb.save('场地信息.xlsx')


def drkb(con):
    '''导入课表'''
    wb = opl.load_workbook(os.path.join(syvar.kbPath, 'kebiao.xlsx'))
    ws = wb.get_sheet_by_name("Sheet1")
    L = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        sksj, qsz, kcdm, kcmc, gzh, jsxm, skddbh, skddmc, lh, xkkh, skbj, kkxy = \
            row[33].value, row[6].value, row[7].value, row[8].value, row[9].value, row[10].value, row[15].value, row[
                16].value, row[28].value, row[20].value, row[19].value, row[23].value
        skbj = skbj.replace(';', ',')
        skddmc = skddmc.replace('理学楼', '理学实验楼')
        t = (sksj, qsz, kcdm, kcmc, gzh, jsxm, skddbh, skddmc, lh, xkkh, skbj, kkxy)
        L.append(t)
    clearTmpTbKb = '''delete from lip_teaching.dbo.jk_gd002_zfjwkb2'''
    con.execute(clearTmpTbKb)
    insertTmpTbkb = '''insert into lip_teaching.dbo.jk_gd002_zfjwkb2(sksj,qsz,kcdm,kcmc,gzh,jsxm,skddbh,skddmc,lh,xkkh,skbj,kkxy)
    values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    '''
    con.execute(insertTmpTbkb, L)
    clearOtherKch = '''
    delete from lip_teaching.dbo.jk_gd002_zfjwkb2 
        where kcdm not in (select kch from lip_baseinfo.dbo.Base_CourseInfo );
    update lip_teaching.dbo.jk_gd002_zfjwkb2 set skddmc=skddmc+'*' where kkxy like '工商%学院';
    update lip_teaching.dbo.jk_gd002_zfjwkb2 set skddmc='公管学院待定地点' where kkxy like '公共%';
    update lip_teaching.dbo.jk_gd002_zfjwkb2 set xs=t.XS,syxs=t.SYZXS from lip_baseinfo.dbo.Base_CourseInfo t
        where lip_teaching.dbo.jk_gd002_zfjwkb2.KCDM=t.kch;
    '''
    con.execute(clearOtherKch)
    clearOtherKch = '''
    delete from lip_teaching.dbo.jk_gd002_zfjwkb2 
        where skddmc not in (select roomname from lip_baseinfo.dbo.Base_RoomInfo);
    delete from lip_teaching.dbo.jk_gd002_zfjwkb;
    insert into lip_teaching.dbo.jk_gd002_zfjwkb select * from lip_teaching.dbo.jk_gd002_zfjwkb2;
    '''
    if syvar.SFJCKCDD:
        con.execute(clearOtherKch)
    else:
        print(
            '请执行 select *from lip_teaching.dbo.jk_gd002_zfjwkb2 where skddmc not in (select roomname from lip_baseinfo.dbo.Base_RoomInfo)，以确保地点正常！')
    return 1


# drkb(connect('mssql'))

def close(con):
    con.close()


def dealSyxm(con):
    '''处理实验项目'''
    print('当前正执行处理实验项目，生成结果位于likai_xmkc')
    first = '''
    if object_id('likai_xmkc') is not null
        begin
        drop table likai_xmkc
        end;
    '''
    con.execute(first)
    createTmpXmKc = '''
    create table lip_baseinfo.dbo.likai_xmkc(
    sssysbh varchar(255),
    syxh varchar(255),
    symc varchar(255),
    syxs varchar(255),
    sylb varchar(255),
    sylx varchar(255),
    syzlb varchar(255),
    xkdm varchar(255),
    syyq varchar(255),
    mzrs varchar(255),
    dgbb varchar(255),
    sskch varchar(255),
    attdepid varchar(255),
    sskc varchar(1000)
    );
    '''
    con.execute(createTmpXmKc)
    # file=os.path.join(syvar.sourcePath,'kcxm.xlsx')
    wb = opl.load_workbook(os.path.join(syvar.sourcePath, 'kcxm.xlsx'))
    ws = wb.active
    L = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        sssysbh, syxh, symc, syxs, sylb, sylx, syzlb, xkdm, syyq, mzrs, dgbb, sskch = \
            row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[
                7].value, row[8].value, row[9].value, row[10].value, row[11].value
        t = (sssysbh, syxh, symc, syxs, sylb, sylx, syzlb, xkdm, syyq, mzrs, dgbb, sskch)
        L.append(t)
    insertTmpXmkc = '''
    insert into lip_baseinfo.dbo.likai_xmkc(sssysbh,syxh,symc,syxs,sylb,sylx,syzlb,xkdm,syyq,mzrs,dgbb,sskch)
        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);
    update lip_baseinfo.dbo.likai_xmkc set lip_baseinfo.dbo.likai_xmkc.attdepid
        =t.id from lip_baseinfo.dbo.Base_Department t where lip_baseinfo.dbo.likai_xmkc.sssysbh=t.DepartmentCode;
    update lip_baseinfo.dbo.likai_xmkc set lip_baseinfo.dbo.likai_xmkc.sskc
        ='('+t.kch+')'+t.KCMC from lip_baseinfo.dbo.Base_CourseInfo t where t.KCH=likai_xmkc.sskch;
    '''
    con.execute(insertTmpXmkc, L)  ###实验项目加载到临时表
    if syvar.SFDCSYXM:
        expSyxm = '''
        select distinct t.sssysbh,t.syxh,t.symc,t.syxs,'' as sylb,t.sylx,t.syzlb,'' as xkdm, '' as syyq,'' as mzrs,t.dgbb,t1.sskch as lastsskch from lip_baseinfo.dbo.likai_xmkc t,
    (select a.syxh,substring(c.sskch,1,len(c.sskch)-1) sskch
    from
    (select distinct SYXH from lip_baseinfo.dbo.likai_xmkc) a
    cross apply
        (select distinct sskch+',' from lip_baseinfo.dbo.likai_xmkc b where a.syxh=b.syxh for xml path('')) c(sskch)) t1
    where t.syxh=t1.syxh 
        '''
        res = con.execute(expSyxm)
        wb = opl.Workbook()
        ws = wb.active
        title = ['所属实验室编号', '实验项目序号', '实验项目名称', '实验学时', '实验类别', '实验类型', '实验者类别', '学科代码',
                 '实验要求', '每组人数', '大纲版本', '所属课程代码']
        ws.append(title)
        for data in res:
            ws.append(data)
        if os.path.exists(os.path.join(syvar.sourcePath, 'lastsyxm.xlsx')):
            os.remove(os.path.join(syvar.sourcePath, 'lastsyxm.xlsx'))
        wb.save(os.path.join(syvar.sourcePath, 'lastsyxm.xlsx'))  ##导出实验项目到本地，符合导出格式的
    return len(L)


def drxm(con, sign='DG'):
    '''直接插入项目进入数据库'''
    print('当前正将项目直接插入至数据库，并且更新信息')
    code = '''
    insert into lip_baseinfo.dbo.Base_Experiment(attdepid,symc,syxs,syzlbid,syzlbmc,Type,UpSYXH)
select
distinct attdepid,symc,syxs,syzlb,'本科生'syzlbmc,'0' type,syxh+'{}' upsyxh
from lip_baseinfo.dbo.likai_xmkc;
--更新实验项目的基本信息
update lip_baseinfo.dbo.Base_Experiment
set SYLBID=t.sylb,SYLXID=t.sylx,SYYQID=t.syyq,MZRS=t.mzrs from lip_baseinfo.dbo.likai_xmkc t
where t.syxh+'{}'=lip_baseinfo.dbo.Base_Experiment.UpSYXH and lip_baseinfo.dbo.Base_Experiment.UpSYXH like '%{}';
--更新实验项目的sskc信息
update lip_baseinfo.dbo.Base_Experiment
set SSKC=t.sskc from
    (select distinct t.syxh+'{}' upsyxh,t1.sskc
    from lip_baseinfo.dbo.likai_xmkc t,
(select a.syxh,substring(c.sskc,1,len(c.sskc)-1) sskc
from
(select distinct syxh from lip_baseinfo.dbo.likai_xmkc) a
cross apply
    (select distinct sskc+',' from lip_baseinfo.dbo.likai_xmkc b where a.syxh=b.syxh for xml path('')) c(sskc)) t1
where t.syxh=t1.syxh) t
where t.upsyxh=lip_baseinfo.dbo.Base_Experiment.UpSYXH and lip_baseinfo.dbo.Base_Experiment.UpSYXH like '%{}';
---建立课程与项目的关系
insert into lip_baseinfo.dbo.Base_Course_Experiment(CourseID, ExperimentID)
select
t2.id courseid ,t3.id experimentid
from lip_baseinfo.dbo.likai_xmkc t1,lip_baseinfo.dbo.Base_CourseInfo t2,lip_baseinfo.dbo.Base_Experiment t3
where t1.sskch=t2.kch and t1.syxh+'{}'=t3.UpSYXH;
--更新实验序号
update lip_baseinfo.dbo.Base_Experiment set SYXH=substring(UpSYXH,7,4)
where SYXH is null and UpSYXH like '%{}';
--更新学科
update lip_baseinfo.dbo.Base_Experiment set xkdmid=t.id from
(select t2.syxh+'{}' upsyxh,t1.id from lip_baseinfo.dbo.Dic_Subject_New t1,lip_baseinfo.dbo.likai_xmkc t2
where t1.SubjectNO=t2.xkdm) t
where t.upsyxh=lip_baseinfo.dbo.Base_Experiment.UpSYXH;
    '''.format(sign, sign, sign, sign, sign, sign, sign, sign)
    con.execute(code)
    return 1

def defaultHzm():
    hzm=datetime.now().strftime('%y')
    return 'DG'+hzm

def syjxxtInterface(mscon,sign):
    print('启动之前，请确保：{}路径下存在kebiao.xlsx文件，如需导入课程，则还需要存在sykc.xlsx,如需要导入项目，则还需要存在kcxm.xlsx'.format(syvar.kbPath))
    if restartService('mssqlserver'):
        # mscon = connect('mssql')
        if restoreMssqlDatabse(mscon):
            print('数据库恢复成功，专业，班级，学生，教师信息更新成功！创建的临时表有:')
            print('zftal_xtgl_jgdmb,zftal_xtgl_zydmb,zftal_xtgl_bjdmb,jw_xjgl_xsjbxxb,view likai_xsbj,likai_jsxx')
        else:
            print('恢复、初始化失败')
        if syvar.SFKQKCTB:
            if drkc(mscon):
                print('课程导入成功，已经进入正式表，创建的临时表有：likai_new_course')
            else:
                print('课程导入失败')
        if drkb(mscon):
            print('课表已导入，已经进入jk_zf_002中')
        else:
            print('课表导入失败')
        if dealSyxm(mscon):
            print('项目处理成功，项目已创建为临时表:')
        else:
            print('项目处理失败')
        if syvar.SFZJINSYXM:
            if not sign or sign.lower().strip()=='default':
                sign=defaultHzm()
            if drxm(mscon,sign):
                print('项目直接导入成功，项目信息更新完成！')
            else:
                print('项目导入失败')
        close(mscon)
        print('处理结束，可以将数据库备份恢复至服务器')
